# -*- coding: utf-8 -*-
import json, re, datetime, tempfile
from pathlib import Path

import streamlit as st
from PIL import Image
import pytesseract
import pdf2image
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

# ─────────────────────────────── Storage helpers ──────────────────────────────
PRESET_PATH = Path("presets.json")

def _default_struct():
    """Initial empty structure."""
    return {"projects": {}}

def load_presets() -> dict:
    if PRESET_PATH.exists():
        try:
            data = json.loads(PRESET_PATH.read_text())
            if "projects" in data and isinstance(data["projects"], dict):
                return data
        except Exception:
            pass
    # fall-back (missing or corrupt)
    PRESET_PATH.write_text(json.dumps(_default_struct(), indent=2))
    return _default_struct()

def save_presets(d: dict):
    PRESET_PATH.write_text(json.dumps(d, indent=2))

presets = load_presets()

# ─────────────────────────────── OCR & Excel helpers ──────────────────────────
BORDER = Border(*(Side(style="thin") for _ in range(4)))

def ocr_text(img: Image.Image, box):
    gray = img.crop(box).convert("L")
    bw = gray.point(lambda x: 0 if x < 180 else 255, "1")
    return pytesseract.image_to_string(bw)

def clean_line(t: str):
    t = (
        t.replace("LOT: STORAGE", "")
        .replace("lJ", "U")
        .replace("l", "I")
        .strip(" :")
    )
    return re.sub(r"\s{2,}", " ", t)

def extract_items(lines):
    """Return [(desc, qty)] with LOT/TYPE."""
    out, i = [], 0
    while i < len(lines):
        m = re.match(r"^(\d+)\s+(.*)", lines[i])
        if m:
            qty, desc = m.group(1), m.group(2)
            if i + 1 < len(lines) and not re.match(r"^\d+\s", lines[i + 1]):
                desc += " " + lines[i + 1]
                i += 1
            if "LOT" in desc.upper() and "TYPE" in desc.upper():
                out.append((clean_line(desc), qty))
        i += 1
    return out

def fill_workbook(tmpl, out, items, meta):
    wb = load_workbook(tmpl)
    ws = wb.active
    header_map = {
        "B5": meta["project"],
        "B6": meta["location"],
        "B7": str(datetime.date.today()),
        "E6": meta["site_contact"],
        "E7": meta["phone"],
    }
    for cell, val in header_map.items():
        if not isinstance(ws[cell], type(ws["A1"]).MergedCell):
            ws[cell].value = val
    r = ws.max_row + 1
    for desc, qty in items:
        for c, val in enumerate((desc, qty, meta["building"], meta["category"]), 1):
            ws.cell(r, c, val).border = BORDER
        r += 1
    wb.save(out)

# ─────────────────────────────── Streamlit config ─────────────────────────────
st.set_page_config(page_title="PDF-to-Excel Loader", layout="wide")
st.title("📑 PDF Shipping-Sheet ➜ Excel Loader")

tab_proc, tab_mgr = st.tabs(["🚚 Process PDF", "🛠️ Preset Manager"])

# ═══════════════════════ 1. Process PDF ═══════════════════════════════════════
with tab_proc:
    if not presets["projects"]:
        st.info("Add a project first in **Preset Manager**.")
    else:
        proj = st.selectbox("Project", list(presets["projects"]))
        ppl = presets["projects"][proj]["personnel"]
        if not ppl:
            st.warning("Add personnel in Preset Manager.")
        else:
            person = st.selectbox("Prepared By", ppl)
            pres_tree = presets["projects"][proj]["presets"]
            if not pres_tree:
                st.warning("Add a preset in Preset Manager.")
            else:
                bldg = st.selectbox("Building", list(pres_tree))
                cats = pres_tree[bldg]
                cat = st.selectbox("Category", list(cats))
                pdf = st.file_uploader("Scanned PDF", ["pdf"])
                xlsx = st.file_uploader("Excel Template", ["xlsx"])
                if st.button("🚀 Generate Excel") and pdf and xlsx:
                    with st.spinner("Running OCR…"):
                        tmp_pdf = Path(tempfile.mktemp(suffix=".pdf"))
                        tmp_pdf.write_bytes(pdf.read())
                        pages = pdf2image.convert_from_path(tmp_pdf)
                        lines = []
                        for pg in pages:
                            w, h = pg.size
                            crop_box = (150, int(h * 0.25), w, int(h * 0.9))
                            lines += [
                                ln.strip()
                                for ln in ocr_text(pg, crop_box).split("\n")
                                if ln.strip()
                            ]
                        items = extract_items(lines)
                        if not items:
                            st.error("No LOT/TYPE rows detected.")
                        else:
                            tmp_xls = Path(tempfile.mktemp(suffix=".xlsx"))
                            tmp_xls.write_bytes(xlsx.read())
                            preset = pres_tree[bldg][cat]
                            meta = {
                                "project": proj,
                                "location": preset["location"],
                                "phone": preset["phone"],
                                "site_contact": preset["contact"],
                                "building": bldg,
                                "category": cat,
                            }
                            fill_workbook(tmp_xls, tmp_xls, items, meta)
                            st.success("Excel ready!")
                            st.download_button(
                                "⬇ Download", tmp_xls.read_bytes(), "filled_template.xlsx"
                            )

# ═══════════════════════ 2. Preset Manager ════════════════════════════════════
with tab_mgr:
    st.subheader("Projects")

    # ── add project
    new_proj = st.text_input("New project name")
    if st.button("Add Project") and new_proj:
        if new_proj in presets["projects"]:
            st.warning("Project exists.")
        else:
            presets["projects"][new_proj] = {"personnel": [], "presets": {}}
            save_presets(presets)
            st.success("Project added."); st.rerun()

    if not presets["projects"]:
        st.stop()

    proj = st.selectbox("Manage project", list(presets["projects"]), key="sel_proj")

    # ── rename / delete project
    colA, colB = st.columns(2)
    with colA:
        new_name = st.text_input("Rename", value=proj, key="rename_proj")
        if st.button("✏ Rename"):
            if new_name and new_name != proj:
                presets["projects"][new_name] = presets["projects"].pop(proj)
                save_presets(presets); st.success("Renamed."); st.rerun()
    with colB:
        if st.button("🗑 Delete Project"):
            presets["projects"].pop(proj)
            save_presets(presets); st.success("Deleted."); st.rerun()

    proj_data = presets["projects"][proj]

    # ── personnel
    st.markdown("### Personnel")
    add_person = st.text_input("Add person")
    if st.button("Add Person") and add_person:
        if add_person not in proj_data["personnel"]:
            proj_data["personnel"].append(add_person)
            save_presets(presets); st.rerun()

    for idx, p in enumerate(proj_data["personnel"]):
        c1, c2, c3 = st.columns([3, 1, 1])
        new_val = c1.text_input(f"pers_{idx}", value=p, key=f"pers_{idx}")
        if c2.button("💾", key=f"save_p_{idx}") and new_val != p:
            proj_data["personnel"][idx] = new_val
            save_presets(presets); st.rerun()
        if c3.button("🗑", key=f"del_p_{idx}"):
            proj_data["personnel"].pop(idx)
            save_presets(presets); st.rerun()

    st.markdown("---\n### Presets")

    # ── add preset form
    with st.form("preset_form", clear_on_submit=True):
        b = st.text_input("Building")
        c = st.text_input("Category")
        loc = st.text_input("Location")
        ph = st.text_input("Phone")
        ct = st.text_input("Contact Name")
        if st.form_submit_button("Save Preset"):
            if all([b, c, loc, ph, ct]):
                proj_data["presets"].setdefault(b, {})[c] = {"location": loc, "phone": ph, "contact": ct}
                save_presets(presets); st.success("Preset saved."); st.rerun()
            else:
                st.warning("Fill all fields.")

    # ── edit / delete existing presets
    edit_state = st.session_state.setdefault("edit_flags", {})
    for bldg, cats in proj_data["presets"].items():
        st.markdown(f"#### 🏢 {bldg}")
        for cat, val in cats.items():
            key = f"{bldg}__{cat}"
            editing = edit_state.get(key, False)
            cols = st.columns([3,3,3,1,1])
            if editing:
                val["location"] = cols[0].text_input("Location", value=val["location"], key=f"loc_{key}")
                val["phone"]    = cols[1].text_input("Phone", value=val["phone"], key=f"ph_{key}")
                val["contact"]  = cols[2].text_input("Contact", value=val["contact"], key=f"ct_{key}")
            else:
                cols[0].markdown(f"**Location:** {val['location']}")
                cols[1].markdown(f"**Phone:** {val['phone']}")
                cols[2].markdown(f"**Contact:** {val['contact']}")

            if not editing:
                if cols[3].button("✏", key=f"edit_{key}"):
                    edit_state[key] = True; st.rerun()
            else:
                if cols[3].button("💾", key=f"save_{key}"):
                    save_presets(presets); edit_state[key] = False; st.rerun()

            if cols[4].button("🗑", key=f"del_{key}"):
                cats.pop(cat)
                if not cats: proj_data["presets"].pop(bldg)
                save_presets(presets); st.rerun()
